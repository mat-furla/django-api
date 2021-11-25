from rest_framework.views import APIView
from django.http import HttpResponse

from users.models import User
from openpyxl import Workbook
from datetime import datetime


class ExportViews(APIView):

    def get(self, req):
        users = User.objects.all()

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Users'

        columns = [
            'id',
            'Email',
            'First Name',
            'Last Name',
            'Birth Day'
        ]
        row_num = 1

        for num, title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=num)
            cell.value = title

        for user in users:
            row_num += 1

            row = [
                user.id,
                user.email,
                user.first_name,
                user.last_name,
                user.birth_date.strftime('%d-%m-%Y')
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        res = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = f'attachment; filename=users.xlsx'
        workbook.save(res)

        return res
